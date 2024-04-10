"""This module is responsible for reading and writing data to/from Excel."""

from io import BytesIO
from datetime import datetime
from dataclasses import dataclass

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(kw_only=True)
class Bilag:
    """A dataclass representing a bilag."""
    sum: float
    text: str
    bilagsart: str
    bilagsnummer: str
    date: datetime


def read_excel(file: BytesIO) -> tuple[Bilag, ...]:
    """Read an Excel sheet and output a list of Bilag objects.

    The columns of the Excel sheet is expected to be:
    SUM, TEKST, AI, BLANK, Bilagsart, Bilagsnummer, FP, Dato, Beløb

    Args:
        file: The Excel file as an BytesIO object.

    Returns:
        A tuple of Bilag objects with length equal to the number of rows in Excel.
    """

    input_sheet: Worksheet = load_workbook(file, read_only=True).active

    bilag_list = []

    iter_ = iter(input_sheet)
    next(iter_)  # Skip header row
    for row in iter_:
        bilagsart = row[4].value

        # Skip rows with bilagsart 'ZF' or None
        if bilagsart == "ZF" or bilagsart is None:
            continue

        bilag = Bilag(
            sum = row[0].value,
            text = row[1].value,
            bilagsart = row[4].value,
            bilagsnummer = row[5].value,
            date = row[7].value
        )
        bilag_list.append(bilag)

    return tuple(bilag_list)


def write_excel(bilag_list: tuple[Bilag, ...], data_list: tuple[tuple[tuple[str, str, float], ...], ...]):
    """Write the given bilag list and data list to an Excel sheet.
    The columns are in the following order:
    SUM, TEKST, Aftale, BLANK, Bilagsart, Bilagsnummer, FP, Dato, Beløb

    Args:
        bilag_list: The list of bilag to write.
        data_list: A list of bilagsdata the same length as the bilag_list.
    """
    wb = Workbook()
    sheet: Worksheet = wb.active

    header = ["SUM", "Tekst", "Aftale", "", "Bilagsart", "Bilagsnummer", "FP", "Dato", "Beløb"]
    sheet.append(header)

    for bilag, data in zip(bilag_list, data_list):
        for postering in data:
            row = [bilag.sum, bilag.text, postering[1], "", bilag.bilagsart, bilag.bilagsnummer, postering[0], bilag.date.date(), postering[2]]
            sheet.append(row)

    file = BytesIO()
    wb.save(file)
    return file
